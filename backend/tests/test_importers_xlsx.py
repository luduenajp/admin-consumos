"""Tests for the Visa XLSX importer."""
from datetime import date
from pathlib import Path

import pytest

from app.importers.visa_xlsx import ParsedPurchaseRow, parse_visa_xlsx


@pytest.fixture()
def minimal_xlsx(tmp_path) -> Path:
    """Create a minimal XLSX file that parse_visa_xlsx can parse."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 1-2: statement metadata
    ws.append(["Fecha de cierre", "Fecha de vencimiento"])
    ws.append(["22/01/2025", "04/02/2025"])
    ws.append([])  # blank row

    # Row 4: header
    ws.append(["Fecha", "Descripción", "Cuotas", "Monto en pesos", "Monto en dólares"])

    # Row 5: a normal purchase
    ws.append(["15/01/2025", "MERPAGO*TIENDA NUBE", "1 de 3", "$10.000,00", ""])

    # Row 6: a USD purchase
    ws.append(["16/01/2025", "STEAM PURCHASE", "", "", "U$S15,99"])

    # Row 7: an excluded row (tax)
    ws.append(["17/01/2025", "DB.RG 5617 Perc.Ganan", "", "$500,00", ""])

    # Row 8: an excluded row (payment)
    ws.append(["18/01/2025", "Su pago en efectivo", "", "$-50.000,00", ""])

    path = tmp_path / "test_statement.xlsx"
    wb.save(path)
    return path


class TestParseVisaXlsx:
    def test_parses_rows(self, minimal_xlsx):
        rows = parse_visa_xlsx(minimal_xlsx)
        # Should parse the normal ARS purchase and the USD purchase, exclude tax and payment
        assert len(rows) == 2

    def test_ars_purchase_parsed_correctly(self, minimal_xlsx):
        rows = parse_visa_xlsx(minimal_xlsx)
        ars_row = [r for r in rows if r.currency == "ARS"][0]
        assert ars_row.purchase_date == date(2025, 1, 15)
        assert ars_row.description == "MERPAGO*TIENDA NUBE"
        assert ars_row.installment_index == 1
        assert ars_row.installments_total == 3
        assert ars_row.installment_amount == 10000.0
        assert ars_row.statement_year_month == "2025-01"

    def test_usd_purchase_parsed_correctly(self, minimal_xlsx):
        rows = parse_visa_xlsx(minimal_xlsx)
        usd_row = [r for r in rows if r.currency == "USD"][0]
        assert usd_row.purchase_date == date(2025, 1, 16)
        assert usd_row.currency == "USD"
        assert usd_row.installment_amount == 15.99

    def test_excluded_descriptions_filtered(self, minimal_xlsx):
        rows = parse_visa_xlsx(minimal_xlsx)
        descriptions = [r.description for r in rows]
        assert all("DB.RG 5617" not in d for d in descriptions)
        assert all("Su pago" not in d for d in descriptions)

    def test_statement_year_month_detected(self, minimal_xlsx):
        rows = parse_visa_xlsx(minimal_xlsx)
        assert all(r.statement_year_month == "2025-01" for r in rows)


class TestParseVisaXlsxInvalid:
    def test_no_statement_date_raises(self, tmp_path):
        """XLSX without 'Fecha de cierre' row should raise ValueError."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Fecha", "Descripción", "Cuotas", "Monto en pesos", "Monto en dólares"])
        ws.append(["15/01/2025", "Test", "", "$1.000,00", ""])

        path = tmp_path / "no_date.xlsx"
        wb.save(path)

        with pytest.raises(ValueError, match="statement year_month"):
            parse_visa_xlsx(path)


class TestOccurrenceTracking:
    def test_duplicate_rows_get_different_occurrence_index(self, tmp_path):
        """Two identical rows on the same day should get occurrence_index 1 and 2."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Fecha de cierre", "Fecha de vencimiento"])
        ws.append(["22/01/2025", "04/02/2025"])
        ws.append([])
        ws.append(["Fecha", "Descripción", "Cuotas", "Monto en pesos", "Monto en dólares"])
        # Two identical rows
        ws.append(["15/01/2025", "SAME PURCHASE", "", "$5.000,00", ""])
        ws.append(["15/01/2025", "SAME PURCHASE", "", "$5.000,00", ""])

        path = tmp_path / "duplicates.xlsx"
        wb.save(path)

        rows = parse_visa_xlsx(path)
        assert len(rows) == 2
        assert rows[0].occurrence_index == 1
        assert rows[1].occurrence_index == 2
