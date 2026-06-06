from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from sqlmodel import Session, select

from app.models import ImportedRow


@dataclass(frozen=True)
class ParsedPurchaseRow:
    purchase_date: date
    description: str
    currency: str  # "ARS" | "USD"
    installment_index: int
    installments_total: int
    installment_amount: float
    statement_year_month: str  # YYYY-MM
    occurrence_index: int = 1
    statement_close_date: Optional[date] = None  # fecha exacta de cierre del resumen


def _parse_ddmmyyyy(value: object) -> Optional[date]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    s = str(value).strip()
    if not s or s.lower() == "nan":
        return None

    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    return None


_money_clean_re = re.compile(r"[^0-9,\-]", re.UNICODE)


def _parse_money(value: object) -> Optional[float]:
    if value is None:
        return None

    s = str(value).strip()
    if not s or s.lower() == "nan":
        return None

    # Examples: "$1.443.685,70", "U$S24,51", "$-7.778,81", "U$S-17,87"
    s = s.replace(".", "")
    s = _money_clean_re.sub("", s)
    if not s:
        return None

    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


_installments_re = re.compile(r"^(\d+)\s*de\s*(\d+)$", re.IGNORECASE)
_installments_c_re = re.compile(r"C\.(\d+)/(\d+)", re.IGNORECASE)  # Banco Nación: C.17/24
_installments_de_re = re.compile(r"(\d+)\s+de\s+(\d+)", re.IGNORECASE)  # MercadoPago: "3 de 3" en medio

_purchase_desc_cleanup_re = re.compile(
    r"\b(?:C\.)\s*\d+\s*/\s*\d+\b|\b\d+\s*de\s*\d+\b",
    re.IGNORECASE,
)

_purchase_desc_leading_code_re = re.compile(r"^\s*\d{3,}\s+")
_purchase_desc_trailing_code_re = re.compile(r"\s+\d{3,}\s*$")


def normalize_purchase_description(*, description: str) -> str:
    cleaned = _purchase_desc_cleanup_re.sub(" ", description)
    cleaned = _purchase_desc_leading_code_re.sub("", cleaned)
    cleaned = _purchase_desc_trailing_code_re.sub("", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def _parse_installments(value: object) -> tuple[int, int]:
    if value is None:
        return 1, 1

    s = str(value).strip()
    if not s or s.lower() == "nan":
        return 1, 1

    m = _installments_re.match(s)
    if not m:
        m = _installments_c_re.search(s)  # C.X/Y en medio del texto
    if not m:
        m = _installments_de_re.search(s)  # "3 de 3" MercadoPago
    if not m:
        return 1, 1

    current = int(m.group(1))
    total = int(m.group(2))
    if total <= 0:
        return 1, 1

    current = max(1, min(current, total))
    return current, total


def _is_excluded_description(description: str) -> bool:
    d = description.strip().lower()
    # Excluir pagos, promos, beneficios, totales
    excluded_prefixes = (
        "su pago",
        "pago de tarjeta",
        "resumen de ",  # "Resumen de enero" (saldo anterior)
        "promo",
        "cr.",
        "cr ",
        "total de",
        "total",
        "tarjeta de",
        "tarjeta visa",
        "movimientos del resumen",
        "bonif.",  # bonificaciones / devoluciones por beneficios
    )
    if d.startswith(excluded_prefixes):
        return True
    # Excluir filas que contengan "total" en cualquier parte
    if "total" in d:
        return True
    # Excluir impuestos (DB.RG 5617, IIBB PERCEP, IMPUESTO DE SELLOS)
    # No excluir devoluciones por compra anulada
    tax_patterns = (
        "db.rg 5617", "iibb percep", "impuesto de sellos", "impuesto de sello", "impuesto al sello",
        "iva rg 4240", "db iva $", "intereses financiacion",
    )
    return any(p in d for p in tax_patterns)


def _detect_statement_year_month(df_raw: pd.DataFrame) -> Optional[str]:
    # Busca la fila donde aparece "Fecha de cierre" y toma el valor de la siguiente fila.
    # En el ejemplo: fila con ["Fecha de cierre", "Fecha de vencimiento"], y luego ["22/01/2026", "04/02/2026"].
    for i in range(len(df_raw) - 1):
        row = df_raw.iloc[i].astype(str).tolist()
        if any("fecha de cierre" in str(v).lower() for v in row):
            next_row = df_raw.iloc[i + 1].tolist()
            for v in next_row:
                d = _parse_ddmmyyyy(v)
                if d is not None:
                    return f"{d.year:04d}-{d.month:02d}"
    return None


def extract_holder_hint_xlsx(
    path: Path,
) -> tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    """
    Extrae (holder_name, last4, card_type, bank) del encabezado de un XLSX de resumen.
    Retorna (None, None, None, None) si no se puede detectar.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None, None, None, None

    holder_name: Optional[str] = None
    last4: Optional[str] = None
    card_type: Optional[str] = None
    bank: Optional[str] = None

    # last4 del nombre del sheet (ej: "Visa 5623")
    sheet_name = wb.sheetnames[0] if wb.sheetnames else ""
    m = re.search(r"(\d{4})$", sheet_name.strip())
    if m:
        last4 = m.group(1)

    # card_type desde el nombre del sheet
    if re.search(r"\bVisa\b", sheet_name, re.IGNORECASE):
        card_type = "Visa"
    elif re.search(r"\bMastercard\b", sheet_name, re.IGNORECASE):
        card_type = "Mastercard"

    ws = wb.worksheets[0]
    for row in ws.iter_rows(max_row=20, values_only=True):
        for cell in row:
            if cell is None:
                continue
            val = str(cell).strip()
            if not val or val.lower() == "none":
                continue

            # "terminada en 5623" → last4
            if last4 is None:
                m = re.search(r"terminada\s+en\s+(\d{4})", val, re.IGNORECASE)
                if m:
                    last4 = m.group(1)

            # card_type desde celdas (ej: "Tarjeta Visa Crédito terminada en...")
            if card_type is None:
                if re.search(r"\bVisa\b", val, re.IGNORECASE):
                    card_type = "Visa"
                elif re.search(r"\bMastercard\b", val, re.IGNORECASE):
                    card_type = "Mastercard"

            # bank: "Movimientos del resumen" es el encabezado del XLSX de Santander
            if bank is None and "movimientos del resumen" in val.lower():
                bank = "santander"

            # bank explícito si aparece nombre de banco
            if bank is None:
                if re.search(r"\bSantander\b", val, re.IGNORECASE):
                    bank = "santander"
                elif re.search(r"\bNacion\b", val, re.IGNORECASE):
                    bank = "nacion"

            # "Juan Pablo Ludueña (Titular)" → holder_name
            if holder_name is None:
                m = re.search(r"^(.+?)\s*\(Titular\)", val, re.IGNORECASE)
                if m:
                    holder_name = m.group(1).strip()

        if holder_name and last4 and card_type and bank:
            break

    wb.close()
    return holder_name, last4, card_type, bank


def parse_visa_xlsx(path: Path) -> list[ParsedPurchaseRow]:
    df_raw = pd.read_excel(path, sheet_name=0, header=None)
    statement_ym = _detect_statement_year_month(df_raw)
    if statement_ym is None:
        raise ValueError("Could not detect statement year_month from XLSX")

    # Detect header row
    header_row_idx: Optional[int] = None
    for i in range(len(df_raw)):
        row = df_raw.iloc[i].astype(str).tolist()
        if any("descripción" in c.lower() for c in row) and any("monto en pesos" in c.lower() for c in row):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Could not find movements table header in XLSX")

    header = df_raw.iloc[header_row_idx].tolist()
    df = df_raw.iloc[header_row_idx + 1 :].copy()
    df.columns = header
    df = df.dropna(how="all")

    out: list[ParsedPurchaseRow] = []
    last_date: Optional[date] = None
    
    # Track occurrences of identical rows in this file to handle multiple same-day identical purchases
    # Key: (date, description, amount, current_installment, total_installments)
    occurrence_tracker: dict[tuple, int] = {}

    for _, r in df.iterrows():
        # Try to get date from current row
        d = _parse_ddmmyyyy(r.get("Fecha"))
        
        # If no date in current row, use the last date we saw
        if d is None:
            if last_date is None:
                continue  # Skip if we don't have any date yet
            d = last_date
        else:
            last_date = d

        description = str(r.get("Descripción") or "").strip()
        if not description or description.lower() == 'nan':
            continue

        if _is_excluded_description(description):
            continue

        amount_ars = _parse_money(r.get("Monto en pesos"))
        amount_usd = _parse_money(r.get("Monto en dólares"))

        currency: Optional[str] = None
        amount: Optional[float] = None
        if amount_ars is not None and amount_ars != 0:
            currency = "ARS"
            amount = amount_ars
        elif amount_usd is not None and amount_usd != 0:
            currency = "USD"
            amount = amount_usd

        if currency is None or amount is None:
            continue

        # Excluir montos negativos (pagos/bonificaciones) por definición de MVP
        if amount <= 0:
            continue

        installment_index, installments_total = _parse_installments(r.get("Cuotas"))
        amount_val = round(float(amount), 2)

        # Update occurrence count
        row_key = (d, description, amount_val, installment_index, installments_total)
        occ_index = occurrence_tracker.get(row_key, 0) + 1
        occurrence_tracker[row_key] = occ_index

        out.append(
            ParsedPurchaseRow(
                purchase_date=d,
                description=description,
                currency=currency,
                installment_index=installment_index,
                installments_total=installments_total,
                installment_amount=amount_val,
                statement_year_month=statement_ym,
                occurrence_index=occ_index,
            )
        )

    return out


def compute_row_fingerprint(*, provider: str, card_id: int, row: ParsedPurchaseRow) -> str:
    payload = {
        "provider": provider,
        "card_id": card_id,
        "purchase_date": row.purchase_date.isoformat(),
        "description": row.description,
        "currency": row.currency,
        "installment_index": row.installment_index,
        "installments_total": row.installments_total,
        "installment_amount": row.installment_amount,
        "statement_year_month": row.statement_year_month,
        "occurrence_index": row.occurrence_index,
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def compute_purchase_fingerprint(*, provider: str, card_id: int, row: ParsedPurchaseRow) -> str:
    payload = {
        "provider": provider,
        "card_id": card_id,
        "purchase_date": row.purchase_date.isoformat(),
        "description": normalize_purchase_description(description=row.description),
        "currency": row.currency,
        "installments_total": row.installments_total,
        "installment_amount": row.installment_amount,
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def was_already_imported(*, session: Session, fingerprint: str) -> bool:
    stmt = select(ImportedRow).where(ImportedRow.row_fingerprint == fingerprint)
    return session.exec(stmt).first() is not None


def mark_imported(*, session: Session, provider: str, source_file: str, fingerprint: str, payload: dict) -> None:
    session.add(
        ImportedRow(
            provider=provider,
            source_file=source_file,
            row_fingerprint=fingerprint,
            parsed_payload_json=json.dumps(payload, ensure_ascii=False),
        )
    )
